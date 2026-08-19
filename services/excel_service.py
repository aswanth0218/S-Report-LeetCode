"""Excel read/write service for S-REPORT."""

from __future__ import annotations

import os
from io import BytesIO
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    CONTEST_RANKING_BUCKETS,
    CONTEST_RATING_BUCKETS,
    EXPORT_FOLDER,
    GLOBAL_RANKING_SUMMARY_BUCKETS,
    INPUT_COLUMNS,
    LEVELS,
    MISSING_DEPT_LABEL,
    OVERALL_CONTEST_RANKING_BUCKETS,
    PROBLEM_SOLVED_CONFIG,
    PROFILE_RANKING_BUCKETS,
    contest_ranking_col,
    global_ranking_col,
    profile_ranking_col,
)


def _save_workbook(wb: Workbook, output_path: str | BytesIO) -> str | BytesIO:
    """Save workbook to a file path or in-memory buffer."""
    wb.save(output_path)
    if isinstance(output_path, BytesIO):
        output_path.seek(0)
    return output_path


def read_input_excel(filepath: str) -> pd.DataFrame:
    """Read uploaded Excel file (.xlsx, .xlsm, or legacy .xls)."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in {".xlsx", ".xls", ".xlsm"}:
        raise ValueError(
            f"Unsupported file type '{ext or 'unknown'}'. Please upload .xlsx or .xls."
        )

    engines: list[str | None]
    if ext == ".xls":
        engines = ["xlrd", None]
    else:
        engines = ["openpyxl", None]

    last_error: Exception | None = None
    for engine in engines:
        try:
            kwargs = {"engine": engine} if engine else {}
            df = pd.read_excel(filepath, **kwargs)
            if df.empty:
                raise ValueError("The Excel file has no data rows.")
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except ImportError as exc:
            last_error = exc
            if ext == ".xls" and engine == "xlrd":
                raise ValueError(
                    "Legacy .xls files are not supported on this server. "
                    "Open the file in Excel and Save As .xlsx, then upload again."
                ) from exc
        except Exception as exc:
            last_error = exc

    if ext == ".xls":
        raise ValueError(
            "Could not read .xls file. Save the workbook as .xlsx in Excel and upload again."
        ) from last_error
    raise ValueError(f"Could not read Excel file: {last_error}") from last_error


def ensure_folders(*folders: str) -> None:
    for folder in folders:
        os.makedirs(folder, exist_ok=True)


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font() -> Font:
    return Font(bold=True, size=11)


def _center_align() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _safe_excel_int(value: Any, default: int = 0) -> int:
    """Convert report counts to int for Excel (handles None/NaN/pd.NA/numpy types)."""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _safe_excel_optional_int(value: Any, default: Any = "") -> Any:
    """Like _safe_excel_int but preserves 0 and uses default for missing values."""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _safe_excel_cell_value(value: Any, default: Any = "") -> Any:
    """Convert any cell value to something openpyxl can write."""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and pd.isna(value):
            return default
        return value
    return value


def _row_value(row: pd.Series, key: str, default: Any = 0) -> Any:
    """Read a value from a report row (handles str/int column names)."""
    if key in row.index:
        val = row[key]
        try:
            if val is None or pd.isna(val):
                return default
        except (TypeError, ValueError):
            if val is None:
                return default
        return val
    if key.isdigit():
        int_key = int(key)
        if int_key in row.index:
            val = row[int_key]
            try:
                if val is None or pd.isna(val):
                    return default
            except (TypeError, ValueError):
                if val is None:
                    return default
            return val
    return default


def _normalize_dept_report_df(dept_report_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Ensure all S-Report columns exist and numeric cells are filled."""
    from config import REPORT_DEPARTMENT_ORDER
    from services.report_service import _init_dept_row

    template_keys = list(_init_dept_row("TEMPLATE").keys())

    if dept_report_df is None or dept_report_df.empty:
        rows = [_init_dept_row(dept) for dept in REPORT_DEPARTMENT_ORDER]
        rows.append(_init_dept_row("Total"))
        return pd.DataFrame(rows)

    df = dept_report_df.copy()
    for key in template_keys:
        if key not in df.columns:
            df[key] = 0

    present_depts = set(df["Dept"].astype(str).tolist()) if "Dept" in df.columns else set()
    missing_dept_rows = [
        _init_dept_row(dept)
        for dept in REPORT_DEPARTMENT_ORDER
        if dept not in present_depts
    ]
    if missing_dept_rows:
        df = pd.concat([df, pd.DataFrame(missing_dept_rows)], ignore_index=True)

    dept_order = REPORT_DEPARTMENT_ORDER + ["Total"]
    ordered_rows: list[dict[str, Any]] = []
    for dept in dept_order:
        match = df[df["Dept"].astype(str) == dept]
        if not match.empty:
            ordered_rows.append(match.iloc[0].to_dict())
    extras = df[~df["Dept"].astype(str).isin(dept_order)]
    if not extras.empty:
        ordered_rows.extend(extras.to_dict(orient="records"))
    df = pd.DataFrame(ordered_rows) if ordered_rows else df

    for key in template_keys:
        if key == "Dept":
            continue
        df[key] = df[key].apply(lambda v: _safe_excel_int(v, 0))

    return df[template_keys]


def generate_s_report_excel(
    dept_report_df: pd.DataFrame,
    missing_dept_df: Optional[pd.DataFrame] = None,
    output_path: Optional[str] = None,
    contest_date: Optional[str] = None,
    contest_title: Optional[str] = None,
) -> str:
    """Generate S-Report.xlsx with grouped headers and formatting."""
    ensure_folders(EXPORT_FOLDER)
    dept_report_df = _normalize_dept_report_df(dept_report_df)

    if output_path is None:
        if contest_date:
            safe_date = contest_date.replace("-", "")
            output_path = os.path.join(EXPORT_FOLDER, f"S-Report_{safe_date}.xlsx")
        else:
            output_path = os.path.join(EXPORT_FOLDER, "S-Report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "S-Report"

    problem_cols = list(PROBLEM_SOLVED_CONFIG.keys())
    ranking_cols = [b["label"] for b in CONTEST_RANKING_BUCKETS]
    rating_cols = [b["label"] for b in CONTEST_RATING_BUCKETS]
    level_cols = LEVELS
    global_ranking_cols = [b["label"] for b in GLOBAL_RANKING_SUMMARY_BUCKETS]

    col_dept = 1
    col_strength = 2
    col_attended = 3
    col_problem_start = 4
    col_ranking_start = col_problem_start + len(problem_cols)
    col_rating_start = col_ranking_start + len(ranking_cols)
    col_level_start = col_rating_start + len(rating_cols)
    col_global_start = col_level_start + len(level_cols)
    total_cols = col_global_start + len(global_ranking_cols) - 1

    header_row_offset = 0
    if contest_date:
        from services.leetcode_service import format_s_report_title

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(
            row=1, column=1,
            value=format_s_report_title(contest_date, contest_title),
        )
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = _center_align()
        header_row_offset = 1

    # Column layout
    # A=Dept, B=Total Strength, C=Total attended
    # D-H: problem solved, I-L: Contest Global Ranking, M-O: rating, P-R: levels, S-V: Overall Global Ranking

    border = _thin_border()
    header_font = _header_font()
    center = _center_align()
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    group_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Row 1 & 2: Group headers (offset when contest date title row present)
    row1 = 1 + header_row_offset
    row2 = 2 + header_row_offset

    ws.cell(row=row1, column=col_dept, value="Dept")
    ws.merge_cells(start_row=row1, start_column=col_dept, end_row=row2, end_column=col_dept)
    ws.cell(row=row1, column=col_strength, value="Total Strength")
    ws.merge_cells(start_row=row1, start_column=col_strength, end_row=row2, end_column=col_strength)
    ws.cell(row=row1, column=col_attended, value="Total attended")
    ws.merge_cells(start_row=row1, start_column=col_attended, end_row=row2, end_column=col_attended)

    ws.cell(row=row1, column=col_problem_start, value="Total Problem solved")
    ws.merge_cells(
        start_row=row1, start_column=col_problem_start,
        end_row=row1, end_column=col_problem_start + len(problem_cols) - 1,
    )
    ws.cell(row=row1, column=col_ranking_start, value="Contest Global Ranking")
    ws.merge_cells(
        start_row=row1, start_column=col_ranking_start,
        end_row=row1, end_column=col_ranking_start + len(ranking_cols) - 1,
    )
    ws.cell(row=row1, column=col_rating_start, value="Contest Rating")
    ws.merge_cells(
        start_row=row1, start_column=col_rating_start,
        end_row=row1, end_column=col_rating_start + len(rating_cols) - 1,
    )
    ws.cell(row=row1, column=col_level_start, value="Levels")
    ws.merge_cells(
        start_row=row1, start_column=col_level_start,
        end_row=row1, end_column=col_level_start + len(level_cols) - 1,
    )
    ws.cell(row=row1, column=col_global_start, value="Overall Global Ranking")
    ws.merge_cells(
        start_row=row1, start_column=col_global_start,
        end_row=row1, end_column=col_global_start + len(global_ranking_cols) - 1,
    )

    for i, col_name in enumerate(problem_cols):
        ws.cell(row=row2, column=col_problem_start + i, value=col_name)
    for i, col_name in enumerate(ranking_cols):
        ws.cell(row=row2, column=col_ranking_start + i, value=col_name)
    for i, col_name in enumerate(rating_cols):
        ws.cell(row=row2, column=col_rating_start + i, value=col_name)
    for i, col_name in enumerate(level_cols):
        ws.cell(row=row2, column=col_level_start + i, value=col_name)
    for i, col_name in enumerate(global_ranking_cols):
        ws.cell(row=row2, column=col_global_start + i, value=col_name)

    for row in (row1, row2):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            cell.fill = group_fill if row == row1 else header_fill

    data_start_row = row2 + 1
    for r_idx, (_, data_row) in enumerate(dept_report_df.iterrows()):
        excel_row = data_start_row + r_idx
        is_total = str(_row_value(data_row, "Dept", "")) == "Total"

        ws.cell(row=excel_row, column=col_dept, value=_safe_excel_cell_value(_row_value(data_row, "Dept", ""), ""))
        ws.cell(row=excel_row, column=col_strength, value=_safe_excel_int(_row_value(data_row, "Total Strength", 0)))
        ws.cell(row=excel_row, column=col_attended, value=_safe_excel_int(_row_value(data_row, "Total attended", 0)))

        for i, key in enumerate(problem_cols):
            ws.cell(row=excel_row, column=col_problem_start + i, value=_safe_excel_int(_row_value(data_row, key, 0)))
        for i, bucket in enumerate(CONTEST_RANKING_BUCKETS):
            key = contest_ranking_col(bucket["label"])
            ws.cell(row=excel_row, column=col_ranking_start + i, value=_safe_excel_int(_row_value(data_row, key, 0)))
        for i, key in enumerate(rating_cols):
            ws.cell(row=excel_row, column=col_rating_start + i, value=_safe_excel_int(_row_value(data_row, key, 0)))
        for i, key in enumerate(level_cols):
            ws.cell(row=excel_row, column=col_level_start + i, value=_safe_excel_int(_row_value(data_row, key, 0)))
        for i, bucket in enumerate(GLOBAL_RANKING_SUMMARY_BUCKETS):
            key = global_ranking_col(bucket["label"])
            ws.cell(row=excel_row, column=col_global_start + i, value=_safe_excel_int(_row_value(data_row, key, 0)))

        for col in range(1, total_cols + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.alignment = center
            cell.border = border
            if is_total:
                cell.font = Font(bold=True, size=11)
                cell.fill = total_fill

    # Column widths
    ws.column_dimensions[get_column_letter(col_dept)].width = 14
    ws.column_dimensions[get_column_letter(col_strength)].width = 16
    ws.column_dimensions[get_column_letter(col_attended)].width = 16
    for col in range(col_problem_start, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    freeze_row = row2 + 1
    ws.freeze_panes = f"A{freeze_row}"

    # Missing Department sheet
    if missing_dept_df is not None and not missing_dept_df.empty:
        ws_missing = wb.create_sheet("Missing Department Students")
        missing_headers = [
            "S.No", "Register No", "Name", "DEPT", "LeetCode Link",
            "Problem Solved", "Contest Rating", "Contest Ranking",
        ]
        for c, header in enumerate(missing_headers, 1):
            cell = ws_missing.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill

        for r_idx, (_, row) in enumerate(missing_dept_df.iterrows(), 2):
            ws_missing.cell(row=r_idx, column=1, value=_safe_excel_optional_int(row.get("S.No")))
            ws_missing.cell(row=r_idx, column=2, value=_safe_excel_optional_int(row.get("Register No")))
            ws_missing.cell(row=r_idx, column=3, value=_safe_excel_cell_value(row.get("Name")))
            ws_missing.cell(row=r_idx, column=4, value=_safe_excel_cell_value(row.get("DEPT")))
            ws_missing.cell(row=r_idx, column=5, value=_safe_excel_cell_value(row.get("LeetCode Link")))
            ws_missing.cell(row=r_idx, column=6, value=_safe_excel_cell_value(row.get("Problem Solved"), 0))
            ws_missing.cell(row=r_idx, column=7, value=_safe_excel_cell_value(row.get("Contest Rating"), ""))
            ws_missing.cell(row=r_idx, column=8, value=_safe_excel_cell_value(row.get("Contest Ranking"), ""))
            for c in range(1, 9):
                ws_missing.cell(row=r_idx, column=c).border = border
                ws_missing.cell(row=r_idx, column=c).alignment = center

        for c in range(1, 9):
            ws_missing.column_dimensions[get_column_letter(c)].width = 18

    return _save_workbook(wb, output_path)


def generate_s_report_excel_bytes(
    dept_report_df: pd.DataFrame,
    missing_dept_df: Optional[pd.DataFrame] = None,
    contest_date: Optional[str] = None,
    contest_title: Optional[str] = None,
) -> BytesIO:
    """Build S-Report.xlsx in memory for direct browser download."""
    buffer = BytesIO()
    generate_s_report_excel(
        dept_report_df,
        missing_dept_df,
        output_path=buffer,
        contest_date=contest_date,
        contest_title=contest_title,
    )
    return buffer


def generate_full_leetcode_excel(
    students_df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Generate Excel with all LeetCode information for every student."""
    from services.format_utils import (
        format_contest_attended,
        format_contest_ranking,
        format_contest_rating,
        format_latest_badge,
        format_badge_details,
        format_profile_ranking,
        format_total_solved,
    )
    from services.leetcode_service import prepare_report_dataframe
    from services.student_data_service import ensure_student_columns

    ensure_folders(EXPORT_FOLDER)
    if output_path is None:
        output_path = os.path.join(EXPORT_FOLDER, "LeetCode-Full-Report.xlsx")

    df = ensure_student_columns(students_df)
    report_df, used_date, contest_title = prepare_report_dataframe(df)

    export_rows = []
    for _, row in report_df.iterrows():
        export_rows.append({
            "S.No": row.get("S.No"),
            "Register No": row.get("Register No"),
            "Name": row.get("Name"),
            "Department (Section)": row.get("Department"),
            "Report Department": row.get("Report Department"),
            "LeetCode Username": row.get("LeetCode Username"),
            "LeetCode Link": row.get("Leetcode Link Clean") or row.get("Leetcode Link"),
            "Fetch Status": row.get("fetch_status"),
            "Total Solved": format_total_solved(row.get("total_solved")),
            "Easy Solved": format_total_solved(row.get("solved_easy")),
            "Medium Solved": format_total_solved(row.get("solved_medium")),
            "Hard Solved": format_total_solved(row.get("solved_hard")),
            "Contest Attended Count": format_contest_attended(row.get("contest_attended")),
            "Contest Rating": format_contest_rating(row.get("contest_rating")),
            "Contest Ranking": format_contest_ranking(row.get("contest_ranking")),
            "Profile Overall Rank": format_profile_ranking(row.get("profile_ranking")),
            "Level": row.get("level"),
            "Award / Badge": format_latest_badge(row.get("latest_badge"), row.get("level")),
            "Badge Details": format_badge_details(row.get("badge_details")),
            "Latest Contest Date": used_date or "",
            "Latest Contest Title": contest_title,
            "Attended Latest Contest": row.get("report_contest_attended", 0),
            "Problems Solved (Latest Contest)": row.get("report_problems_solved", 0),
            "Rating (Latest Contest)": format_contest_rating(row.get("report_contest_rating")),
            "Ranking (Latest Contest)": format_contest_ranking(row.get("report_contest_ranking")),
        })

    export_df = pd.DataFrame(export_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "All Students"
    _write_dataframe_sheet(ws, export_df)

    summary_df = _build_overall_department_summary(report_df)
    ws2 = wb.create_sheet("Department Summary")
    _write_dataframe_sheet(ws2, summary_df)

    return _save_workbook(wb, output_path)


def generate_solved_problems_excel(
    students_df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Generate Excel with solved counts and topic breakdown per difficulty."""
    from services.format_utils import format_total_solved
    from services.leetcode_service import SOLVED_PROBLEMS_LANGUAGE_COLUMNS, build_topic_report_fields
    from services.student_data_service import ensure_student_columns

    ensure_folders(EXPORT_FOLDER)
    if output_path is None:
        output_path = os.path.join(EXPORT_FOLDER, "Solved-Problems-Report.xlsx")

    df = ensure_student_columns(students_df)
    export_rows = []
    for _, row in df.iterrows():
        topic_fields = build_topic_report_fields({
            "easy_topics": row.get("easy_topics") or [],
            "medium_topics": row.get("medium_topics") or [],
            "hard_topics": row.get("hard_topics") or [],
            "solved_languages": row.get("solved_languages") or [],
        })
        easy_text = row.get("easy_topics_text") or topic_fields["easy_topics_text"]
        medium_text = row.get("medium_topics_text") or topic_fields["medium_topics_text"]
        hard_text = row.get("hard_topics_text") or topic_fields["hard_topics_text"]
        languages_text = row.get("solved_languages_text") or topic_fields["solved_languages_text"]

        export_rows.append({
            "S.No": row.get("S.No"),
            "Name": row.get("Name"),
            "Register No": row.get("Register No"),
            "Department": row.get("Department") or row.get("DEPT") or "-",
            "Overall Problems Solved": format_total_solved(row.get("total_solved")),
            "Easy Count": format_total_solved(row.get("solved_easy")),
            "Easy Topics": easy_text,
            "Medium Count": format_total_solved(row.get("solved_medium")),
            "Medium Topics": medium_text,
            "Hard Count": format_total_solved(row.get("solved_hard")),
            "Hard Topics": hard_text,
            **{label: topic_fields.get(label, 0) for label in SOLVED_PROBLEMS_LANGUAGE_COLUMNS},
            "Languages Used": languages_text,
        })

    export_df = pd.DataFrame(export_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Solved Problems"
    _write_solved_problems_sheet(ws, export_df)
    return _save_workbook(wb, output_path)


def _write_solved_problems_sheet(ws, export_df: pd.DataFrame) -> None:
    border = _thin_border()
    header_font = _header_font()
    center = _center_align()
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    topic_columns = {"Easy Topics", "Medium Topics", "Hard Topics", "Languages Used"}

    headers = list(export_df.columns)
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill
        width = 48 if header in topic_columns else max(14, len(str(header)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = width

    for r_idx, row in enumerate(export_df.itertuples(index=False), 2):
        for c, value in enumerate(row, 1):
            header = headers[c - 1]
            cell = ws.cell(row=r_idx, column=c, value=_safe_excel_cell_value(value))
            cell.border = border
            cell.alignment = wrap_left if header in topic_columns else center

    ws.freeze_panes = "A2"


def _safe_num(value: Any) -> float:
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _badge_category(latest_badge: Any, level: Any = None) -> str:
    from services.format_utils import format_latest_badge

    label = format_latest_badge(latest_badge, level)
    if label == "Guardian":
        return "Guardian"
    if label == "Knight" or "knight" in label.lower():
        return "Knight"
    if label == "-":
        return "Unrated"
    return "Other"


def _build_overall_profile_rows(students_df: pd.DataFrame) -> pd.DataFrame:
    """Student-level overall profile rows for department report export."""
    from services.format_utils import (
        format_contest_attended,
        format_latest_badge,
        format_badge_details,
        format_profile_ranking,
        format_total_solved,
    )
    from services.student_data_service import ensure_student_columns

    df = ensure_student_columns(students_df)
    rows = []
    for _, row in df.iterrows():
        rows.append({
            "S.No": row.get("S.No"),
            "Register No": row.get("Register No"),
            "Name": row.get("Name"),
            "Department": row.get("Department"),
            "Report Department": row.get("Report Department"),
            "LeetCode Username": row.get("LeetCode Username"),
            "Total Solved": format_total_solved(row.get("total_solved")),
            "Contests Attended": format_contest_attended(row.get("contest_attended")),
            "Profile Rank": format_profile_ranking(row.get("profile_ranking")),
            "Level": row.get("level"),
            "Award / Badge": format_latest_badge(row.get("latest_badge"), row.get("level")),
            "Badge Details": format_badge_details(row.get("badge_details")),
        })
    return pd.DataFrame(rows)


def _build_overall_department_summary(students_df: pd.DataFrame) -> pd.DataFrame:
    """Build grouped department summary with LeetCode totals."""
    from config import MISSING_DEPT_LABEL, REPORT_DEPARTMENT_ORDER
    from services.student_data_service import ensure_student_columns

    df = ensure_student_columns(students_df)
    valid = df[df["Report Department"] != MISSING_DEPT_LABEL]

    rows = []
    for dept in REPORT_DEPARTMENT_ORDER:
        dept_students = valid[valid["Report Department"] == dept]
        strength = len(dept_students)
        ranks = [_safe_num(v) for v in dept_students["profile_ranking"] if _safe_num(v) > 0]
        badge_labels = dept_students.apply(
            lambda r: _badge_category(r.get("latest_badge"), r.get("level")),
            axis=1,
        )
        rows.append({
            "Dept": dept,
            "Total Strength": strength,
            "Total Solved (Sum)": int(dept_students["total_solved"].fillna(0).sum()),
            "Avg Total Solved": round(dept_students["total_solved"].fillna(0).mean(), 1) if strength else 0,
            "Easy (Sum)": int(dept_students["solved_easy"].fillna(0).sum()),
            "Medium (Sum)": int(dept_students["solved_medium"].fillna(0).sum()),
            "Hard (Sum)": int(dept_students["solved_hard"].fillna(0).sum()),
            "Contest Attended Count": int((dept_students["contest_attended"].fillna(0) > 0).sum()),
            "Guardian": int((badge_labels == "Guardian").sum()),
            "Knight": int((badge_labels == "Knight").sum()),
            "Other Awards": int((badge_labels == "Other").sum()),
            "Unrated": int((badge_labels == "Unrated").sum()),
            "Avg Profile Rank": int(round(sum(ranks) / len(ranks))) if ranks else "-",
        })

    total_badges = valid.apply(
        lambda r: _badge_category(r.get("latest_badge"), r.get("level")),
        axis=1,
    )
    total = {
        "Dept": "Total",
        "Total Strength": sum(r["Total Strength"] for r in rows),
        "Total Solved (Sum)": sum(r["Total Solved (Sum)"] for r in rows),
        "Avg Total Solved": round(
            pd.to_numeric(valid["total_solved"], errors="coerce").fillna(0).mean(), 1
        ) if len(valid) else 0,
        "Easy (Sum)": sum(r["Easy (Sum)"] for r in rows),
        "Medium (Sum)": sum(r["Medium (Sum)"] for r in rows),
        "Hard (Sum)": sum(r["Hard (Sum)"] for r in rows),
        "Contest Attended Count": sum(r["Contest Attended Count"] for r in rows),
        "Guardian": int((total_badges == "Guardian").sum()),
        "Knight": int((total_badges == "Knight").sum()),
        "Other Awards": int((total_badges == "Other").sum()),
        "Unrated": int((total_badges == "Unrated").sum()),
        "Avg Profile Rank": "-",
    }
    rows.append(total)
    return pd.DataFrame(rows)


def _write_dataframe_sheet(ws, export_df: pd.DataFrame) -> None:
    border = _thin_border()
    header_font = _header_font()
    center = _center_align()
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    headers = list(export_df.columns)
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(str(header)) + 2)

    for r_idx, row in enumerate(export_df.itertuples(index=False), 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c, value=_safe_excel_cell_value(value))
            cell.border = border
            cell.alignment = center

    ws.freeze_panes = "A2"


def generate_weekly_contest_details_excel(
    students_df: pd.DataFrame,
    contest_date: Optional[str] = None,
    contest_filter: str = "",
    department_filter: str = "",
    s_no_filter: str = "",
    register_filter: str = "",
    name_filter: str = "",
    search_filter: str = "",
    link_status_filter: str = "",
    problems_filter: str = "",
    rating_filter: str = "",
    rank_filter: str = "",
    global_rank_filter: str = "",
    sort_by: Optional[str] = None,
    sort_dir: str = "asc",
    output_path: str | BytesIO | None = None,
) -> str | BytesIO:
    """Export weekly contest details for all students to Excel."""
    from services.leetcode_service import format_s_report_title
    from services.report_service import get_weekly_contest_details

    ensure_folders(EXPORT_FOLDER)
    if output_path is None:
        output_path = os.path.join(EXPORT_FOLDER, "Weekly-Contest-Details.xlsx")

    records, used_date, contest_title, _ = get_weekly_contest_details(
        students_df,
        contest_date=contest_date,
        contest_filter=contest_filter,
        department_filter=department_filter,
        s_no_filter=s_no_filter,
        register_filter=register_filter,
        name_filter=name_filter,
        link_status_filter=link_status_filter,
        problems_filter=problems_filter,
        rating_filter=rating_filter,
        rank_filter=rank_filter,
        global_rank_filter=global_rank_filter,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    export_columns = [
        "S.No",
        "Register No",
        "Student Name",
        "Department",
        "Contest Rating",
        "Level",
        "Contest Rank",
        "Solved",
        "Overall Global Ranking",
        "Total Problems Solved",
        "Easy",
        "Medium",
        "Hard",
        "Profile Rank",
    ]
    if records:
        raw_df = pd.DataFrame(records)
        if "raw_global_ranking" in raw_df.columns:
            raw_df["Overall Global Ranking"] = raw_df["raw_global_ranking"].apply(
                lambda v: int(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() not in ("-", "", "N/A", "nan") else "N/A"
            )
        elif "Global Ranking" in raw_df.columns and "Overall Global Ranking" not in raw_df.columns:
            raw_df["Overall Global Ranking"] = raw_df["Global Ranking"]

        if "Student Name" not in raw_df.columns and "Name" in raw_df.columns:
            raw_df["Student Name"] = raw_df["Name"]
        if "Solved" not in raw_df.columns and "Problems Solved (out of 4)" in raw_df.columns:
            raw_df["Solved"] = raw_df["Problems Solved (out of 4)"]
        if "Level" not in raw_df.columns and "level" in raw_df.columns:
            raw_df["Level"] = raw_df["level"]
        if "Total Problems Solved" not in raw_df.columns:
            for alt in ("Total Problem Solved", "Totall problem solved", "Overall Problems Solved", "total_solved"):
                if alt in raw_df.columns:
                    raw_df["Total Problems Solved"] = raw_df[alt]
                    break
        if "Profile Rank" not in raw_df.columns:
            for alt in ("profile_rank", "profile_ranking", "Profile ranking"):
                if alt in raw_df.columns:
                    raw_df["Profile Rank"] = raw_df[alt]
                    break

        export_df = raw_df[[col for col in export_columns if col in raw_df.columns]]
    else:
        export_df = pd.DataFrame(columns=export_columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Contest"
    if used_date:
        title = format_s_report_title(used_date, contest_title)
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_df.columns))
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        start_row = 2
    else:
        start_row = 1

    headers = list(export_df.columns)
    header_font = _header_font()
    center = _center_align()
    border = _thin_border()
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(str(header)) + 2)

    for r_idx, row in enumerate(export_df.itertuples(index=False), start_row + 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c, value=_safe_excel_cell_value(value))
            cell.border = border
            cell.alignment = center

    ws.freeze_panes = f"A{start_row + 1}"
    if len(headers) > 0 and len(export_df) > 0:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(export_df)}"
    return _save_workbook(wb, output_path)


def generate_student_contest_report_excel(
    students_df: pd.DataFrame,
    contest_date: Optional[str] = None,
    output_path: str | BytesIO | None = None,
    department: Optional[str] = None,
) -> str | BytesIO:
    """Export student-wise weekly contest and profile summary to Excel."""
    from services.leetcode_service import format_s_report_title
    from services.report_service import get_student_contest_report_records

    ensure_folders(EXPORT_FOLDER)
    if output_path is None:
        output_path = os.path.join(EXPORT_FOLDER, "Student-Contest-Report.xlsx")

    records, used_date, contest_title = get_student_contest_report_records(
        students_df,
        contest_date=contest_date,
        department=department,
    )
    export_columns = [
        "S.No",
        "Register Number",
        "Name",
        "Department",
        "Profile URL",
        "Contest (Attended / Not Attended)",
        "Problems Solved (out of 4)",
        "Contest Rating",
        "Contest Rank (Weekly Contest Rank)",
        "Overall Global Ranking (Overall / Best)",
        "Overall Problems Solved",
        "Easy",
        "Medium",
        "Hard",
    ]
    if records:
        raw_df = pd.DataFrame(records)
        if "Contest Rank (Weekly Contest Rank)" not in raw_df.columns and "Contest Rank" in raw_df.columns:
            raw_df["Contest Rank (Weekly Contest Rank)"] = raw_df["Contest Rank"]
        if "Overall Global Ranking (Overall / Best)" not in raw_df.columns:
            for alt_col in ("Overall Global Ranking", "Global Profile Ranking", "Global Ranking"):
                if alt_col in raw_df.columns:
                    raw_df["Overall Global Ranking (Overall / Best)"] = raw_df[alt_col]
                    break
        export_df = raw_df[[col for col in export_columns if col in raw_df.columns]]
    else:
        export_df = pd.DataFrame(columns=export_columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Contest Report"
    if used_date:
        title = format_s_report_title(used_date, contest_title)
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_df.columns))
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        start_row = 2
    else:
        start_row = 1

    headers = list(export_df.columns)
    header_font = _header_font()
    center = _center_align()
    border = _thin_border()
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill
        width = 18 if header == "Profile URL" else max(14, len(str(header)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = width

    for r_idx, row in enumerate(export_df.itertuples(index=False), start_row + 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c, value=_safe_excel_cell_value(value))
            cell.border = border
            cell.alignment = center if c != 5 else Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = f"A{start_row + 1}"
    if len(headers) > 0 and len(export_df) > 0:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(export_df)}"
    return _save_workbook(wb, output_path)


def generate_overall_department_excel(
    students_df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Generate overall department LeetCode summary Excel."""
    ensure_folders(EXPORT_FOLDER)
    if output_path is None:
        output_path = os.path.join(EXPORT_FOLDER, "Overall-Department-Report.xlsx")

    from services.student_data_service import ensure_student_columns

    df = ensure_student_columns(students_df)
    summary_df = _build_overall_department_summary(df)
    profile_df = _build_overall_profile_rows(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Overall Department"
    _write_dataframe_sheet(ws, summary_df)

    ws_profile = wb.create_sheet("Overall Profile")
    _write_dataframe_sheet(ws_profile, profile_df)

    return _save_workbook(wb, output_path)


def _sample_input_dataframe() -> pd.DataFrame:
    """Build the sample upload workbook used for Download Sample Input."""
    data = {
        "S.No": [1, 2, 3, 4, 5, 6, 7, 8],
        "Register No": [
            "732124205001", "732124205002", "732124205003", "732124205004",
            "732124205005", "732124205006", "732124205007", "732124205008",
        ],
        "Name": [
            "AAKIL SHIHAB S", "ABINAYA S", "ABI SRI S", "ABIVARSHA N",
            "TEST STUDENT IT", "SAMPLE ECE", "SAMPLE EEE", "NO DEPT",
        ],
        "DEPT": ["CSE A", "CSE B", "AI & DS", "IT A", "IT B", "ECE", "EEE", ""],
        "Leetcode Link": [
            "[https://leetcode.com/u/Aakil-shihab14/](https://leetcode.com/u/Aakil-shihab14/)",
            "https://leetcode.com/u/Abinaya1910/",
            "https://leetcode.com/u/_Abisri_/",
            "https://leetcode.com/u/abivarsha/",
            "https://leetcode.com/u/invalid_user_xyz12345/",
            "https://leetcode.com/u/invalid_user_xyz12345/",
            "https://leetcode.com/u/invalid_user_xyz12345/",
            "",
        ],
    }
    return pd.DataFrame(data)


def create_sample_input_excel_bytes() -> BytesIO:
    """Create sample input Excel in memory for direct download."""
    buffer = BytesIO()
    _sample_input_dataframe().to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return buffer


def create_sample_input_excel(output_path: str) -> str:
    """Create a sample input Excel file for testing."""
    ensure_folders(os.path.dirname(output_path) or ".")
    _sample_input_dataframe().to_excel(output_path, index=False, engine="openpyxl")
    return output_path


def generate_missing_data_issues_excel(
    students_df: pd.DataFrame,
    output_path: str | BytesIO | None = None,
) -> str | BytesIO:
    """Generate comprehensive Excel workbook containing all data issues for faculty correction."""
    from services.student_data_service import ensure_student_columns
    from services.validation_service import (
        _is_blank,
        get_leetcode_profile_link_status,
        is_fetchable_leetcode_profile,
    )

    ensure_folders(EXPORT_FOLDER)
    if output_path is None:
        output_path = os.path.join(EXPORT_FOLDER, "Missing-Data-Report.xlsx")

    df = ensure_student_columns(students_df) if students_df is not None else pd.DataFrame()

    all_issues: list[dict[str, Any]] = []
    missing_dept: list[dict[str, Any]] = []
    link_issues: list[dict[str, Any]] = []
    duplicate_regs: list[dict[str, Any]] = []

    reg_seen: dict[str, list[dict[str, Any]]] = {}

    for _, row in df.iterrows():
        sno = row.get("S.No")
        reg_no = row.get("Register No")
        name = row.get("Name")
        dept = row.get("DEPT") or row.get("Department")
        raw_dept = str(dept).strip() if dept is not None and not pd.isna(dept) else ""
        link = row.get("Leetcode Link") or row.get("LeetCode Link")
        raw_link = str(link).strip() if link is not None and not pd.isna(link) else ""
        fetch_status = str(row.get("fetch_status") or "").strip()

        row_issues: list[str] = []
        fix_suggestions: list[str] = []

        # Check Department
        if (
            _is_blank(dept)
            or raw_dept.upper() in ("MISSING DEPARTMENT", "<NA>", "NAN", "NONE", "BLANK", "")
            or dept == MISSING_DEPT_LABEL
            or row.get("Report Department") == MISSING_DEPT_LABEL
        ):
            row_issues.append("Missing Department")
            fix_suggestions.append("Enter department code (e.g. CSE A, IT B, AI & DS, ECE, EEE)")
            missing_dept.append({
                "S.No": sno,
                "Register No": reg_no,
                "Name": name,
                "DEPT": raw_dept or "BLANK",
                "Leetcode Link": raw_link,
                "Issue": "Missing Department",
                "Action Needed": "Provide valid department",
            })

        # Check Register No
        if _is_blank(reg_no):
            row_issues.append("Missing Register No")
            fix_suggestions.append("Enter valid register number")
        else:
            reg_key = str(reg_no).strip()
            if reg_key not in reg_seen:
                reg_seen[reg_key] = []
            reg_seen[reg_key].append({
                "S.No": sno, "Register No": reg_no, "Name": name, "DEPT": raw_dept,
            })

        # Check Name
        if _is_blank(name):
            row_issues.append("Missing Name")
            fix_suggestions.append("Enter student full name")

        # Check LeetCode Link
        if _is_blank(link):
            row_issues.append("Missing LeetCode Link")
            fix_suggestions.append("Provide profile URL (e.g. https://leetcode.com/u/username/)")
            link_issues.append({
                "S.No": sno, "Register No": reg_no, "Name": name, "DEPT": raw_dept,
                "Leetcode Link": "BLANK", "Issue": "Missing Link",
                "Action Needed": "Add valid LeetCode profile URL",
            })
        elif not is_fetchable_leetcode_profile(link):
            row_issues.append("Invalid LeetCode Link / Placeholder")
            fix_suggestions.append("Replace placeholder/invalid text with valid LeetCode profile URL")
            link_issues.append({
                "S.No": sno, "Register No": reg_no, "Name": name, "DEPT": raw_dept,
                "Leetcode Link": raw_link, "Issue": "Invalid URL / Placeholder",
                "Action Needed": "Provide valid LeetCode profile URL",
            })
        elif fetch_status in ("Profile Not Found", "Profile Fetch Failed", "Fetch Failed"):
            row_issues.append(f"LeetCode Fetch Status: {fetch_status}")
            fix_suggestions.append("Check if LeetCode username exists or is spelled correctly")
            link_issues.append({
                "S.No": sno, "Register No": reg_no, "Name": name, "DEPT": raw_dept,
                "Leetcode Link": raw_link, "Issue": fetch_status,
                "Action Needed": "Verify username exists on LeetCode",
            })

        if row_issues:
            all_issues.append({
                "S.No": sno,
                "Register No": reg_no,
                "Name": name,
                "DEPT": raw_dept,
                "Leetcode Link": raw_link,
                "Issue Type": "; ".join(row_issues),
                "Suggested Fix": "; ".join(fix_suggestions),
            })

    # Track duplicates
    for reg_key, entries in reg_seen.items():
        if len(entries) > 1:
            for item in entries:
                duplicate_regs.append({
                    "S.No": item.get("S.No"),
                    "Register No": item.get("Register No"),
                    "Name": item.get("Name"),
                    "DEPT": item.get("DEPT"),
                    "Issue": f"Duplicate Register No (appears {len(entries)} times)",
                })

    wb = Workbook()
    border = _thin_border()
    header_font = _header_font()
    center = _center_align()
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def _write_sheet(ws_obj, title: str, records: list[dict[str, Any]], fallback_cols: list[str]) -> None:
        ws_obj.title = title
        cols = list(records[0].keys()) if records else fallback_cols
        for c, h in enumerate(cols, 1):
            cell = ws_obj.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill

        if records:
            for r_idx, rec in enumerate(records, 2):
                for c, key in enumerate(cols, 1):
                    val = _safe_excel_cell_value(rec.get(key))
                    cell = ws_obj.cell(row=r_idx, column=c, value=val)
                    cell.border = border
                    if key in ("Name", "Leetcode Link", "Issue Type", "Suggested Fix", "Action Needed"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = center
        else:
            ws_obj.cell(row=2, column=1, value="No issues found in this category.")
            ws_obj.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(cols)))

        for c, col_name in enumerate(cols, 1):
            width = max(14, len(str(col_name)) + 4)
            if col_name in ("Leetcode Link", "Issue Type", "Suggested Fix"):
                width = 32
            elif col_name in ("Name", "DEPT"):
                width = 22
            ws_obj.column_dimensions[get_column_letter(c)].width = width
        ws_obj.freeze_panes = "A2"

    ws_all = wb.active
    _write_sheet(
        ws_all,
        "All Data Issues",
        all_issues,
        ["S.No", "Register No", "Name", "DEPT", "Leetcode Link", "Issue Type", "Suggested Fix"],
    )

    ws_dept = wb.create_sheet("Missing Department")
    _write_sheet(
        ws_dept,
        "Missing Department",
        missing_dept,
        ["S.No", "Register No", "Name", "DEPT", "Leetcode Link", "Issue", "Action Needed"],
    )

    ws_links = wb.create_sheet("Link Issues")
    _write_sheet(
        ws_links,
        "Link Issues",
        link_issues,
        ["S.No", "Register No", "Name", "DEPT", "Leetcode Link", "Issue", "Action Needed"],
    )

    if duplicate_regs:
        ws_dups = wb.create_sheet("Duplicate Register Nos")
        _write_sheet(
            ws_dups,
            "Duplicate Register Nos",
            duplicate_regs,
            ["S.No", "Register No", "Name", "DEPT", "Issue"],
        )

    return _save_workbook(wb, output_path)


def generate_missing_data_issues_excel_bytes(students_df: pd.DataFrame) -> BytesIO:
    """Create Missing Data Issues Excel workbook in memory for instant browser download."""
    buffer = BytesIO()
    generate_missing_data_issues_excel(students_df, output_path=buffer)
    buffer.seek(0)
    return buffer
