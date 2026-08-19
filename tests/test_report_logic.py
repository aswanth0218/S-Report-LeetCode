"""Basic tests for report calculation logic."""

import pandas as pd

from services.report_service import calculate_grouped_department_report, classify_problem_solved


def test_classify_problem_solved_contest_range():
    assert classify_problem_solved(4) == "4"
    assert classify_problem_solved(3) == "3"
    assert classify_problem_solved(0) == "0"
    assert classify_problem_solved(500) == "4"  # clamped to 4 max for weekly contest


def test_grouped_report_problem_buckets_sum_to_strength():
    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "contest_history": [{
                "contest_date": "2026-01-01", "attended": True,
                "problems_solved": 4, "rating": 1600.0, "ranking": 3000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "B", "DEPT": "CSE B",
            "Department": "CSE B", "Report Department": "CSE",
            "contest_history": [{
                "contest_date": "2026-01-01", "attended": True,
                "problems_solved": 2, "rating": 1400.0, "ranking": 8000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 3, "Register No": "R3", "Name": "C", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    report_df, _, quality, _, _ = calculate_grouped_department_report(
        students, contest_date="2026-01-01"
    )
    cse = report_df[report_df["Dept"] == "CSE"].iloc[0]

    assert cse["Total Strength"] == 3
    assert cse["Total attended"] == 2
    assert cse["4"] == 1
    assert cse["2"] == 1
    assert cse["0"] == 0
    assert cse["4"] + cse["3"] + cse["2"] + cse["1"] + cse["0"] == cse["Total attended"]
    assert cse["4"] <= cse["Total attended"]
    assert cse["Guardian"] + cse["Knight"] + cse["Unrated"] == cse["Total attended"]

    assert quality["passed"] is True


def test_contest_lookup_excel_format():
    """Contest lookup Excel (Reg No, Student Name, Department, Solved out of 4)."""
    import pandas as pd

    from services.report_service import calculate_grouped_department_report
    from services.validation_service import prepare_students_dataframe

    raw = pd.DataFrame([
        {
            "S.No": 1,
            "Reg No": 732123104002,
            "Department": "CSE A",
            "Student Name": "Ajay E",
            "Contest Rating": 1655,
            "Level": "Knight",
            "Contest Rank": 6157,
            "Solved (out of 4)": "3 / 4",
        },
        {
            "S.No": 2,
            "Reg No": 732123104003,
            "Department": "IT B",
            "Student Name": "Priya S",
            "Contest Rating": 1400,
            "Level": "Unrated",
            "Contest Rank": 12000,
            "Solved (out of 4)": "2 / 4",
        },
    ])

    students = prepare_students_dataframe(raw, contest_date="2026-08-02")
    assert students.iloc[0]["Department"] == "CSE A"
    assert students.iloc[0]["Report Department"] == "CSE"
    assert students.iloc[0]["excel_problems_solved"] == 3
    assert students.iloc[0]["level"] == "Knight"

    report_df, _, quality, used_date, _ = calculate_grouped_department_report(
        students, contest_date="2026-08-02"
    )
    cse = report_df[report_df["Dept"] == "CSE"].iloc[0]
    it = report_df[report_df["Dept"] == "IT"].iloc[0]
    total = report_df[report_df["Dept"] == "Total"].iloc[0]

    assert used_date == "2026-08-02"
    assert cse["Total Strength"] == 1
    assert cse["Total attended"] == 1
    assert cse["3"] == 1
    assert it["2"] == 1
    assert total["Total Strength"] == 2
    assert total["Total attended"] == 2
    assert cse["Knight"] == 1
    assert total["Knight"] + total["Unrated"] == total["Total attended"]
    assert quality["passed"] is True


def test_problem_buckets_only_count_attendees():
    """Non-attendees must not inflate the 0-solved bucket."""
    students = pd.DataFrame([
        {
            "S.No": i, "Register No": f"R{i}", "Name": f"S{i}",
            "DEPT": "CSE A", "Department": "CSE A", "Report Department": "CSE",
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": ps, "rating": 1500.0, "ranking": 6000,
            }] if attended else [],
            "level": "Unrated",
        }
        for i, (attended, ps) in enumerate([
            (True, 3), (True, 2), (True, 1), (True, 0), (True, 0),
            (False, 0), (False, 0), (False, 0),
        ], start=1)
    ])

    report_df, _, quality, _, _ = calculate_grouped_department_report(
        students, contest_date="2026-08-02"
    )
    cse = report_df[report_df["Dept"] == "CSE"].iloc[0]

    assert cse["Total Strength"] == 8
    assert cse["Total attended"] == 5
    assert cse["3"] == 1
    assert cse["2"] == 1
    assert cse["1"] == 1
    assert cse["0"] == 2
    assert cse["4"] + cse["3"] + cse["2"] + cse["1"] + cse["0"] == 5
    assert cse["Below 1500"] == 5
    assert cse["Guardian"] + cse["Knight"] + cse["Unrated"] == 5
    assert quality["passed"] is True


def test_rating_and_ranking_boundary_values():
    from services.leetcode_service import classify_contest_rating, classify_contest_ranking

    assert classify_contest_rating(1499) == "Below 1500"
    assert classify_contest_rating(1500) == "Below 1500"
    assert classify_contest_rating(1501) == "1501 - 2000"
    assert classify_contest_ranking(4999) == "Below 5000"
    assert classify_contest_ranking(5000) == "Below 5000"
    assert classify_contest_ranking(5001) == "5001 - 10000"
    assert classify_contest_ranking(10001) == "10001 - 15000"
    assert classify_contest_ranking(15001) == "Above 15000"
    from services.leetcode_service import classify_profile_ranking
    assert classify_profile_ranking(19999) == "Below 20000"
    assert classify_profile_ranking(20000) == "Below 20000"
    assert classify_profile_ranking(20001) == "20000 < 100000"
    assert classify_profile_ranking(100000) == "20000 < 100000"
    assert classify_profile_ranking(100001) == "Above 100000"


def test_leetcode_link_cleaning_and_usernames():
    from services.leetcode_service import (
        clean_leetcode_link,
        extract_username,
        is_fetchable_leetcode_profile,
    )
    from services.validation_service import prepare_students_dataframe

    assert extract_username("https://leetcode.com/u/alice/") == "alice"
    assert extract_username("leetcode.com/u/bob") == "bob"
    assert extract_username("www.leetcode.com/u/carol/") == "carol"
    assert extract_username("dave_123") == "dave_123"
    assert extract_username("Long Absent") is None
    assert extract_username("NA") is None
    assert is_fetchable_leetcode_profile("Long Absent") is False
    assert is_fetchable_leetcode_profile("leetcode.com/u/eve") is True

    cleaned = clean_leetcode_link("leetcode.com/u/frank/")
    assert cleaned == "https://leetcode.com/u/frank/"

    df = pd.DataFrame([{
        "S.No": 1,
        "Register No": 732124205001,
        "Name": "Test",
        "DEPT": "CSE A",
        "Leetcode Link": "leetcode.com/u/test_user",
    }])
    prepared = prepare_students_dataframe(df)
    assert prepared.iloc[0]["LeetCode Username"] == "test_user"
    assert bool(prepared.iloc[0]["Has Valid Link"]) is True
    assert int(prepared.iloc[0]["Register No"]) == 732124205001


def test_eee_department_filter_with_blank_department_column():
    from services.report_service import compute_student_filter_counts, filter_students

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "E1",
            "DEPT": "EEE", "Department": "EEE", "contest_attended": 1,
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "E2",
            "DEPT": "EEE A", "Department": "EEE A", "contest_attended": 0,
        },
        {
            "S.No": 3, "Register No": "R3", "Name": "E3",
            "DEPT": "EEE", "Department": "", "contest_attended": 1,
        },
        {
            "S.No": 4, "Register No": "R4", "Name": "E4",
            "DEPT": "EEE B", "Department": None, "contest_attended": 0,
        },
        {
            "S.No": 5, "Register No": "R5", "Name": "C1",
            "DEPT": "ECE", "Department": "ECE", "contest_attended": 1,
        },
    ])

    counts = compute_student_filter_counts(students, {})
    assert counts["departments"]["EEE"] == 4
    assert counts["departments"]["ECE"] == 1

    filtered = filter_students(students, {"department": "EEE"})
    assert len(filtered) == 4
    assert set(filtered["Name"].tolist()) == {"E1", "E2", "E3", "E4"}


def test_department_filter_slug_and_sections():
    from services.report_service import filter_students, normalize_department_filter

    assert normalize_department_filter("AI_DS") == "AI&DS"

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "Department": "CSE A",
            "Report Department": "CSE", "contest_attended": 1,
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "B", "Department": "IT",
            "Report Department": "IT", "contest_attended": 0,
        },
    ])

    assert len(filter_students(students, {"department": "CSE"})) == 1
    assert len(filter_students(students, {"department": "AI_DS"})) == 0
    assert len(filter_students(students, {"contest": "attended"})) == 1


def test_extract_latest_leetcode_badge():
    from services.leetcode_service import extract_badge_details, extract_latest_leetcode_badge

    profile = {
        "success": True,
        "raw": {
            "matchedUser": {
                "badges": [
                    {"displayName": "500 Days Badge", "creationDate": 1000},
                    {"displayName": "Guardian", "creationDate": 2000},
                ],
                "contestBadge": {"displayName": "Knight"},
            }
        },
    }
    assert extract_latest_leetcode_badge(profile) == "Guardian"
    details = extract_badge_details(profile)
    assert "Guardian" in details
    assert "500 Days Badge" in details
    assert details.index("Guardian") < details.index("500 Days Badge")

    contest_only = {
        "success": True,
        "raw": {"matchedUser": {"badges": [], "contestBadge": {"displayName": "Knight"}}},
    }
    assert extract_latest_leetcode_badge(contest_only) == "Knight"
    assert extract_badge_details(contest_only) == "Knight (Contest)"


def test_student_details_lifetime_contest_count():
    """Student Details Contests column shows lifetime contests attended, not weekly 0/1."""
    from services.report_service import filter_students, prepare_student_details_dataframe

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "Department": "CSE A",
            "Report Department": "CSE", "LeetCode Username": "alice",
            "contest_attended": 0,
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 3, "rating": 1500.0, "ranking": 5000,
            }],
            "level": "Knight",
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "B", "Department": "CSE B",
            "Report Department": "CSE", "LeetCode Username": "bob",
            "contest_attended": 12,
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    details_df, used_date, _ = prepare_student_details_dataframe(students, contest_date="2026-08-02")
    assert used_date == "2026-08-02"
    assert int(details_df.iloc[0]["contest_attended"]) == 0
    assert int(details_df.iloc[1]["contest_attended"]) == 12
    assert int(details_df.iloc[0]["report_contest_attended"]) == 1

    search_hit = filter_students(details_df, {"search": "R2"})
    assert len(search_hit) == 1
    assert search_hit.iloc[0]["Register No"] == "R2" or str(search_hit.iloc[0]["Register No"]) == "R2"


def test_filter_and_sort_students():
    from services.report_service import filter_students, sort_students

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R001", "Name": "Alice", "Department": "CSE A",
            "Report Department": "CSE", "LeetCode Username": "alice1",
            "total_solved": 10, "solved_easy": 5, "solved_medium": 3, "solved_hard": 2,
            "contest_attended": 2, "contest_rating": 1600, "contest_ranking": 4000,
            "profile_ranking": 100000, "level": "Knight",
        },
        {
            "S.No": 2, "Register No": "R002", "Name": "Bob", "Department": "IT",
            "Report Department": "IT", "LeetCode Username": "bob2",
            "total_solved": 3, "solved_easy": 2, "solved_medium": 1, "solved_hard": 0,
            "contest_attended": 0, "contest_rating": None, "contest_ranking": None,
            "profile_ranking": 500000, "level": "Unrated",
        },
    ])

    assert len(filter_students(students, {"department": "CSE"})) == 1

    low_to_high = sort_students(students, "total_solved", "asc")
    assert list(low_to_high["total_solved"]) == [3, 10]

    high_to_low = sort_students(students, "total_solved", "desc")
    assert list(high_to_low["total_solved"]) == [10, 3]

    by_name = sort_students(students, "name", "asc")
    assert list(by_name["Name"]) == ["Alice", "Bob"]

    from services.report_service import get_student_sort_columns
    sort_cols = get_student_sort_columns()
    profile_rank_col = next((c for c in sort_cols if c["param"] == "profile_ranking"), None)
    assert profile_rank_col is not None
    assert profile_rank_col["label"] == "Profile Rank"


def test_sort_students_by_badge_rank():
    from services.format_utils import badge_sort_rank
    from services.report_service import sort_students

    students = pd.DataFrame([
        {
            "S.No": 1, "Name": "A", "latest_badge": "-", "level": "Unrated",
            "badge_details": "-",
        },
        {
            "S.No": 2, "Name": "B", "latest_badge": "Knight", "level": "Knight",
            "badge_details": "Knight (2026-01-01)",
        },
        {
            "S.No": 3, "Name": "C", "latest_badge": "Guardian", "level": "Guardian",
            "badge_details": "Guardian (2026-02-01)",
        },
    ])

    assert badge_sort_rank("Guardian", "Guardian") > badge_sort_rank("Knight", "Knight")
    high_to_low = sort_students(students, "badge", "desc")
    assert list(high_to_low["Name"]) == ["C", "B", "A"]


def test_parse_badge_display_list_splits_badges():
    from services.format_utils import format_student_for_display, parse_badge_display_list

    details = "Guardian (2026-02-01); 500 Days Badge (2026-01-01); Knight (Contest)"
    assert parse_badge_display_list(details) == [
        "Guardian (2026-02-01)",
        "500 Days Badge (2026-01-01)",
        "Knight (Contest)",
    ]

    formatted = format_student_for_display({
        "badge_details": details,
        "latest_badge": "Guardian",
        "level": "Guardian",
    })
    assert len(formatted["badge_items"]) == 3
    assert formatted["badge_items"][0]["name"] == "Guardian (2026-02-01)"


def test_all_contest_buckets_sum_to_attended():
    """Every attendee must appear in exactly one bucket per contest category."""
    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 3, "rating": 1500.0, "ranking": 5000,
            }],
            "level": "Knight",
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "B", "DEPT": "IT A",
            "Department": "IT A", "Report Department": "IT",
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 1, "rating": 1650.0, "ranking": 8000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 3, "Register No": "R3", "Name": "C", "DEPT": "IT B",
            "Department": "IT B", "Report Department": "IT",
            "excel_contest_attended": 1,
            "excel_problems_solved": 2,
            "excel_contest_rating": 1400.0,
            "level": "Unrated",
        },
        {
            "S.No": 4, "Register No": "R4", "Name": "D", "DEPT": "ECE",
            "Department": "ECE", "Report Department": "ECE",
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    report_df, _, quality, _, _ = calculate_grouped_department_report(
        students, contest_date="2026-08-02"
    )
    total = report_df[report_df["Dept"] == "Total"].iloc[0]
    attended = int(total["Total attended"])

    assert attended == 3
    assert sum(int(total[k]) for k in ("4", "3", "2", "1", "0")) == attended
    from config import CONTEST_RANKING_BUCKETS, CONTEST_RATING_BUCKETS, PROFILE_RANKING_BUCKETS
    from config import contest_ranking_col, profile_ranking_col
    assert sum(int(total[contest_ranking_col(b["label"])]) for b in CONTEST_RANKING_BUCKETS) == attended
    assert sum(int(total[b["label"]]) for b in CONTEST_RATING_BUCKETS) == attended
    assert int(total["Guardian"]) + int(total["Knight"]) + int(total["Unrated"]) == attended
    assert sum(int(total[profile_ranking_col(b["label"])]) for b in PROFILE_RANKING_BUCKETS) == int(total["Total Strength"])
    assert quality["passed"] is True


def test_profile_and_contest_ranking_buckets_do_not_collide():
    """Profile 10001-15000 and contest 10001-15000 must use separate counters."""
    from config import PROFILE_RANKING_BUCKETS, contest_ranking_col, profile_ranking_col

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "profile_ranking": 5000001,
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 2, "rating": 1600.0, "ranking": 12000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "B", "DEPT": "CSE B",
            "Department": "CSE B", "Report Department": "CSE",
            "profile_ranking": 8000,
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    report_df, _, quality, _, _ = calculate_grouped_department_report(
        students, contest_date="2026-08-02"
    )
    cse = report_df[report_df["Dept"] == "CSE"].iloc[0]
    total = report_df[report_df["Dept"] == "Total"].iloc[0]

    assert cse["Total Strength"] == 2
    assert cse["Total attended"] == 1
    assert cse[profile_ranking_col("Above 100000")] == 1
    assert cse[profile_ranking_col("Below 20000")] == 1
    assert cse[contest_ranking_col("10001 - 15000")] == 1
    assert cse[profile_ranking_col("20000 < 100000")] == 0
    assert sum(int(total[profile_ranking_col(b["label"])]) for b in PROFILE_RANKING_BUCKETS) == int(total["Total Strength"])
    assert quality["passed"] is True


def test_latest_weekly_contest_uses_student_attended_date():
    """Latest report should use the most recent contest students actually attended."""
    from services.leetcode_service import prepare_report_dataframe, resolve_report_contest_date

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "contest_history": [{
                "contest_date": "2026-07-19", "attended": True,
                "problems_solved": 2, "rating": 1600.0, "ranking": 12000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 2, "Register No": "R2", "Name": "B", "DEPT": "IT A",
            "Department": "IT A", "Report Department": "IT",
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    assert resolve_report_contest_date(students, contest_date=None) == "2026-07-19"

    report_df, used_date, _ = prepare_report_dataframe(students, contest_date=None)
    assert used_date == "2026-07-19"
    assert int(report_df.iloc[0]["report_contest_attended"]) == 1
    assert int(report_df.iloc[0]["report_problems_solved"]) == 2
    assert int(report_df.iloc[1]["report_contest_attended"]) == 0

    report_df2, _, _, _, _ = calculate_grouped_department_report(students, contest_date=None)
    total = report_df2[report_df2["Dept"] == "Total"].iloc[0]
    assert int(total["Total attended"]) == 1
    assert int(total["2"]) == 1


def test_weekly_contest_details_filters_and_records():
    from services.report_service import get_weekly_contest_details

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": 101, "Name": "Alice", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "Leetcode Link": "https://leetcode.com/u/alice/",
            "Has Valid Link": True, "fetch_status": "Success",
            "profile_ranking": 259135,
            "contest_ranking": 8500,
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 3, "rating": 1650.0, "ranking": 9000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 2, "Register No": 102, "Name": "Bob", "DEPT": "EEE A",
            "Department": "EEE A", "Report Department": "EEE",
            "Leetcode Link": "https://leetcode.com/u/bob/",
            "Has Valid Link": True, "fetch_status": "Profile Not Found",
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    all_records, used_date, _, counts = get_weekly_contest_details(
        students, contest_date="2026-08-02"
    )
    assert used_date == "2026-08-02"
    assert counts["link_status"]["Success"] == 1
    assert counts["link_status"]["Profile Not Found"] == 1
    assert counts["contest"] == {"all": 2, "attended": 1, "not_attended": 1}
    assert all_records[0]["Department"] in ("CSE", "CSE A")
    assert all_records[0]["Leetcode profile Link Status"] == "Success"
    assert all_records[0]["Contest"] == "Attended"
    assert all_records[0]["Contest Rank"] == 9000
    assert all_records[0]["Global Ranking"] == "8,500"
    assert all_records[0]["raw_global_ranking"] == 8500
    assert all_records[1]["Contest"] == "Not Attended"
    assert all_records[1]["Leetcode profile Link Status"] == "Profile Not Found"
    assert all_records[1]["Problems Solved (out of 4)"] == "-"
    assert all_records[1]["Global Ranking"] == "N/A"

    attended_only, _, _, attended_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", contest_filter="attended", department_filter="CSE"
    )
    assert attended_counts["contest"]["attended"] == 1
    assert len(attended_only) == 1
    assert attended_only[0]["Register No"] == 101

    by_problems, _, _, problem_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", problems_filter="3"
    )
    assert problem_counts["problems_solved"]["3"] == 1
    assert len(by_problems) == 1
    assert by_problems[0]["Problems Solved (out of 4)"] == "3"

    by_rating, _, _, rating_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", rating_filter="1501 - 2000"
    )
    assert rating_counts["contest_rating"]["1501 - 2000"] == 1
    assert len(by_rating) == 1

    by_rank, _, _, rank_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", rank_filter="5001 - 10000"
    )
    assert rank_counts["contest_rank"]["5001 - 10000"] == 1
    assert len(by_rank) == 1

    by_global_rank, _, _, global_rank_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", global_rank_filter="Below 20,000"
    )
    assert len(by_global_rank) == 1
    assert by_global_rank[0]["Global Ranking"] == "8,500"

    by_na_global_rank, _, _, _ = get_weekly_contest_details(
        students, contest_date="2026-08-02", global_rank_filter="N/A / Not Available"
    )
    assert len(by_na_global_rank) == 1
    assert by_na_global_rank[0]["Name"] == "Bob"

    by_link, _, _, link_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", link_status_filter="Profile Not Found"
    )
    assert link_counts["link_status"]["Profile Not Found"] == 1
    assert len(by_link) == 1
    assert by_link[0]["Register No"] == 102

    by_dept, _, _, dept_counts = get_weekly_contest_details(
        students, contest_date="2026-08-02", department_filter="CSE"
    )
    assert dept_counts["department"]["CSE"] == 1
    assert len(by_dept) == 1

    by_name, _, _, _ = get_weekly_contest_details(
        students, contest_date="2026-08-02", name_filter="alice"
    )
    assert len(by_name) == 1
    assert by_name[0]["Name"] == "Alice"

    sorted_asc, _, _, _ = get_weekly_contest_details(
        students, contest_date="2026-08-02", sort_by="global_ranking", sort_dir="asc"
    )
    assert sorted_asc[0]["raw_global_ranking"] == 8500

    sorted_desc, _, _, _ = get_weekly_contest_details(
        students,
        contest_date="2026-08-02",
        sort_by="global_ranking",
        sort_dir="desc",
        name_filter="alice",
    )
    assert sorted_desc[0]["raw_global_ranking"] == 8500

def test_global_ranking_best_selection_and_filters():
    from services.report_service import (
        get_best_global_ranking,
        matches_global_ranking_filter,
        get_weekly_contest_details,
    )

    # 1. Best ranking selects contest_ranking from profile contest stats
    row_multi = pd.Series({
        "contest_ranking": 15420,
        "profile_ranking": 8210,
        "excel_contest_ranking": 12500,
    })
    assert get_best_global_ranking(row_multi) == 15420

    # 2. Student with only profile_ranking returns None because left-side profile_ranking must not be used for contest Global Ranking
    row_profile_only = pd.Series({
        "contest_ranking": None,
        "profile_ranking": 8210,
        "excel_contest_ranking": None,
    })
    assert get_best_global_ranking(row_profile_only) is None

    # 3. Missing/invalid rankings return None
    row_na = pd.Series({"contest_ranking": None, "profile_ranking": "N/A"})
    assert get_best_global_ranking(row_na) is None

    # 3. Filter boundary tests
    assert matches_global_ranking_filter(2087, "Below 20000") is True
    assert matches_global_ranking_filter(19999, "Below 20000") is True
    assert matches_global_ranking_filter(20000, "Below 20000") is True
    assert matches_global_ranking_filter(20001, "Below 20000") is False

    assert matches_global_ranking_filter(20000, "20000 - 100000") is True
    assert matches_global_ranking_filter(45500, "20000 - 100000") is True
    assert matches_global_ranking_filter(100000, "20000 - 100000") is True
    assert matches_global_ranking_filter(100001, "20000 - 100000") is False

    assert matches_global_ranking_filter(100001, "Above 100000") is True
    assert matches_global_ranking_filter(500000, "Above 100000") is True
    assert matches_global_ranking_filter(100000, "Above 100000") is False

    assert matches_global_ranking_filter(None, "N/A / Not Available") is True
    assert matches_global_ranking_filter(0, "N/A / Not Available") is True
    assert matches_global_ranking_filter(2087, "N/A / Not Available") is False


def test_department_report_display_includes_global_ranking():
    from services.report_service import get_department_report_display_columns

    columns = get_department_report_display_columns()
    assert "contest_ranking" in columns
    labels = [label for label, _ in columns["contest_ranking"]]
    assert labels == ["Below 5000", "5001 - 10000", "10001 - 15000", "Above 15000"]
    assert "global_ranking" in columns
    global_labels = [label for label, _ in columns["global_ranking"]]
    assert global_labels == ["Below 20000", "20000 < 100000", "Above 100000", "N/A"]


def test_department_summary_global_ranking_aggregation():
    from config import global_ranking_col
    from services.report_service import calculate_grouped_department_report

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": 101, "Name": "Alice", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "contest_ranking": 2087, "excel_contest_attended": 1,
            "excel_problems_solved": 4, "excel_contest_rating": 2109, "level": "Guardian",
        },
        {
            "S.No": 2, "Register No": 102, "Name": "Bob", "DEPT": "CSE B",
            "Department": "CSE B", "Report Department": "CSE",
            "contest_ranking": 45000, "excel_contest_attended": 1,
            "excel_problems_solved": 2, "excel_contest_rating": 1700, "level": "Knight",
        },
        {
            "S.No": 3, "Register No": 103, "Name": "Charlie", "DEPT": "IT A",
            "Department": "IT A", "Report Department": "IT",
            "contest_ranking": 150000, "excel_contest_attended": 1,
            "excel_problems_solved": 1, "excel_contest_rating": 1400, "level": "Unrated",
        },
        {
            "S.No": 4, "Register No": 104, "Name": "David", "DEPT": "ECE A",
            "Department": "ECE A", "Report Department": "ECE",
            "contest_ranking": None, "excel_contest_attended": 0,
            "excel_problems_solved": 0, "excel_contest_rating": 1500, "level": "Unrated",
        },
    ])

    report_df, _, _, _, _ = calculate_grouped_department_report(students)
    cse_row = report_df[report_df["Dept"] == "CSE"].iloc[0]
    assert cse_row["Total Strength"] == 2
    assert cse_row[global_ranking_col("Below 20000")] == 1
    assert cse_row[global_ranking_col("20000 < 100000")] == 1
    assert cse_row[global_ranking_col("Above 100000")] == 0
    assert cse_row[global_ranking_col("N/A")] == 0

    it_row = report_df[report_df["Dept"] == "IT"].iloc[0]
    assert it_row["Total Strength"] == 1
    assert it_row[global_ranking_col("Above 100000")] == 1

    ece_row = report_df[report_df["Dept"] == "ECE"].iloc[0]
    assert ece_row["Total Strength"] == 1
    assert ece_row[global_ranking_col("N/A")] == 1

    total_row = report_df[report_df["Dept"] == "Total"].iloc[0]
    assert total_row["Total Strength"] == 4
    assert total_row[global_ranking_col("Below 20000")] == 1
    assert total_row[global_ranking_col("20000 < 100000")] == 1
    assert total_row[global_ranking_col("Above 100000")] == 1
    assert total_row[global_ranking_col("N/A")] == 1

    # Check filtering by global rank
    filtered_report, _, _, _, _ = calculate_grouped_department_report(
        students, global_rank_filter="Below 20000"
    )
    cse_filtered = filtered_report[filtered_report["Dept"] == "CSE"].iloc[0]
    assert cse_filtered["Total Strength"] == 1
    tot_filtered = filtered_report[filtered_report["Dept"] == "Total"].iloc[0]
    assert tot_filtered["Total Strength"] == 1
    assert tot_filtered[global_ranking_col("Below 20000")] == 1


def test_student_contest_report_records():
    from services.report_service import get_student_contest_report_records

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": 101, "Name": "Alice", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "Leetcode Link": "https://leetcode.com/u/alice/",
            "Leetcode Link Clean": "https://leetcode.com/u/alice/",
            "total_solved": 120, "solved_easy": 40, "solved_medium": 50, "solved_hard": 30,
            "profile_ranking": 259135,
            "contest_ranking": 8500,
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 3, "rating": 1650.0, "ranking": 9000,
            }],
            "level": "Knight",
        },
        {
            "S.No": 2, "Register No": 102, "Name": "Bob", "DEPT": "EEE A",
            "Department": "EEE A", "Report Department": "EEE",
            "Leetcode Link": "https://leetcode.com/u/bob/",
            "total_solved": 10, "solved_easy": 5, "solved_medium": 3, "solved_hard": 2,
            "profile_ranking": 350000,
            "contest_ranking": 50000,
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    records, used_date, _ = get_student_contest_report_records(students, contest_date="2026-08-02")
    assert used_date == "2026-08-02"
    assert len(records) == 2
    assert records[0]["Contest (Attended / Not Attended)"] == "Attended"
    assert records[0]["Problems Solved (out of 4)"] == "3"
    assert records[0]["Department"] == "CSE"
    assert records[0]["Contest Rank"] == 9000
    assert records[0]["Contest Rank (Weekly Contest Rank)"] == 9000
    assert records[0]["Overall Global Ranking"] == "8,500"
    assert records[0]["Overall Contest Global Ranking"] == "8,500"
    assert records[0]["Easy"] == 40
    assert records[0]["Medium"] == 50
    assert records[0]["Hard"] == 30
    assert records[1]["Contest (Attended / Not Attended)"] == "Not Attended"
    assert records[1]["Problems Solved (out of 4)"] == "-"
    assert records[1]["Overall Global Ranking"] == "50,000"
    assert records[1]["Overall Contest Global Ranking"] == "50,000"


def test_build_language_solved_counts():
    from services.leetcode_service import build_language_solved_counts

    languages = [
        {"languageName": "Java", "problemsSolved": 12},
        {"languageName": "Python", "problemsSolved": 8},
        {"languageName": "Python3", "problemsSolved": 5},
        {"languageName": "C", "problemsSolved": 3},
        {"languageName": "C++", "problemsSolved": 15},
        {"languageName": "MySQL", "problemsSolved": 2},
    ]
    counts = build_language_solved_counts(languages)
    assert counts == {
        "Java": 12,
        "Python": 8,
        "Python3": 5,
        "C": 3,
        "C++": 15,
        "MySQL": 2,
    }


def test_parse_contest_date_param_latest_vs_custom():
    from services.leetcode_service import parse_contest_date_param

    assert parse_contest_date_param("", "") is None
    assert parse_contest_date_param("", "2026-08-02") == "2026-08-02"
    assert parse_contest_date_param("2026-07-19", "2026-08-02") == "2026-07-19"


def test_s_report_handles_pandas_na_in_missing_dept():
    import tempfile

    from openpyxl import load_workbook

    from services.excel_service import generate_s_report_excel
    from services.report_service import calculate_grouped_department_report

    students = pd.DataFrame([
        {
            "S.No": pd.NA, "Register No": pd.NA, "Name": "Missing Dept Student",
            "DEPT": "", "Department": "Missing Department", "Report Department": "Missing Department",
            "contest_history": [], "level": "Unrated",
        },
        {
            "S.No": 1, "Register No": 732124205001, "Name": "Valid",
            "DEPT": "CSE A", "Department": "CSE A", "Report Department": "CSE",
            "contest_history": [{
                "contest_date": "2026-07-19", "attended": True,
                "problems_solved": 2, "rating": 1600.0, "ranking": 12000,
            }],
            "level": "Unrated",
        },
    ])

    report_df, missing_df, _, used_date, _ = calculate_grouped_department_report(students, contest_date=None)
    assert int(report_df[report_df["Dept"] == "Total"]["Total attended"].iloc[0]) == 1

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        generate_s_report_excel(report_df, missing_df, output_path=path, contest_date=used_date)
        wb = load_workbook(path)
        assert "Missing Department Students" in wb.sheetnames
    finally:
        import os
        os.remove(path)


def test_undated_excel_contest_data_used_for_latest_report():
    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "excel_contest_attended": 1,
            "excel_problems_solved": 3,
            "excel_contest_rating": 1650.0,
            "excel_contest_ranking": 8000,
            "contest_history": [],
            "level": "Unrated",
        },
    ])

    report_df, _, quality, used_date, _ = calculate_grouped_department_report(students, contest_date=None)
    total = report_df[report_df["Dept"] == "Total"].iloc[0]
    assert used_date
    assert int(total["Total attended"]) == 1
    assert int(total["3"]) == 1
    assert quality["passed"] is True


def test_read_input_excel_sample_xlsx():
    import os

    from services.excel_service import create_sample_input_excel, read_input_excel

    sample_path = os.path.join(os.path.dirname(__file__), "..", "uploads", "test_sample_read.xlsx")
    os.makedirs(os.path.dirname(sample_path), exist_ok=True)
    create_sample_input_excel(sample_path)
    try:
        df = read_input_excel(sample_path)
        assert len(df) == 8
        assert "Register No" in df.columns or "Leetcode Link" in df.columns
    finally:
        if os.path.exists(sample_path):
            os.remove(sample_path)


def test_format_s_report_title():
    from services.leetcode_service import format_s_report_title

    assert format_s_report_title("2026-08-02", "Weekly Contest 511") == (
        "Weekly Contest 511 Report — 2026-08-02"
    )
    assert format_s_report_title("2026-08-02", "Weekly Contest (2026-08-02)") == (
        "Weekly Contest Report — 2026-08-02"
    )


def test_s_report_excel_writes_data(tmp_path=None):
    """S-Report Excel must contain department rows with numeric values."""
    import os
    import tempfile

    from openpyxl import load_workbook

    from services.excel_service import generate_s_report_excel
    from services.report_service import calculate_grouped_department_report

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "R1", "Name": "A", "DEPT": "CSE A",
            "Department": "CSE A", "Report Department": "CSE",
            "excel_contest_attended": 1,
            "excel_problems_solved": 3,
            "excel_contest_rating": 1655.0,
            "excel_contest_ranking": 6157,
            "level": "Knight",
        },
    ])

    report_df, missing_df, _, used_date, title = calculate_grouped_department_report(
        students, contest_date="2026-08-02"
    )
    out_dir = tempfile.mkdtemp()
    output_path = os.path.join(out_dir, "S-Report_test.xlsx")
    generate_s_report_excel(
        report_df, missing_df,
        output_path=output_path,
        contest_date=used_date,
        contest_title=title,
    )

    wb = load_workbook(output_path)
    ws = wb["S-Report"]
    # Data starts after title + 2 header rows
    data_row = 4
    assert ws.cell(row=data_row, column=1).value == "CSE"
    assert _safe_int(ws.cell(row=data_row, column=2).value) == 1
    assert _safe_int(ws.cell(row=data_row, column=3).value) == 1
    assert _safe_int(ws.cell(row=data_row, column=5).value) == 1  # problem bucket "3"

    # Header check: Levels is the last group header in S-Report.xlsx
    # A=1, B=2, C=3, D-H(5)=4-8, I-L(4)=9-12, M-O(3)=13-15, P-R(3)=16-18
    assert ws.cell(row=2, column=16).value == "Levels"
    assert ws.cell(row=3, column=16).value == "Guardian"
    assert ws.cell(row=3, column=17).value == "Knight"
    assert ws.cell(row=3, column=18).value == "Unrated"


def test_all_report_departments_group_correctly():
    """Every standard report department groups section variants and appears in report."""
    from services.department_service import get_report_department
    from services.student_data_service import ensure_student_columns

    expected = {
        "CSE": ("CSE A", "CSE B", "CSE-A"),
        "AI&DS": ("AIDS", "AI&DS A", "AIDS B", "AI & DS", "AI-DS"),
        "IT": ("IT A", "IT B", "IT-A"),
        "EEE": ("EEE A", "EEE B", "EE", "E & E"),
        "ECE": ("ECE A", "ECE B", "ECE-A"),
    }
    for report_dept, variants in expected.items():
        for variant in variants:
            assert get_report_department(variant) == report_dept, variant

    students = []
    contest = {
        "contest_date": "2026-08-02", "attended": True,
        "problems_solved": 1, "rating": 1500.0, "ranking": 7000,
    }
    dept_variants = ["CSE A", "AIDS A", "IT B", "EEE A", "ECE B"]
    for idx, dept in enumerate(dept_variants, start=1):
        students.append({
            "S.No": idx, "DEPT": dept, "Department": dept,
            "Report Department": dept, "Name": f"Student {idx}",
            "contest_history": [contest], "level": "Unrated",
        })

    fixed = ensure_student_columns(pd.DataFrame(students))
    assert set(fixed["Report Department"]) == {"CSE", "AI&DS", "IT", "EEE", "ECE"}

    report_df, _, _, _, _ = calculate_grouped_department_report(
        fixed, contest_date="2026-08-02"
    )
    report_depts = [row for row in report_df["Dept"].tolist() if row != "Total"]
    assert report_depts == ["CSE", "AI&DS", "IT", "ECE", "EEE"]
    for dept in report_depts:
        row = report_df[report_df["Dept"] == dept].iloc[0]
        assert row["Total Strength"] == 1
        assert row["Total attended"] == 1


def test_eee_section_departments_group_into_eee():
    from services.department_service import get_report_department
    from services.student_data_service import ensure_student_columns

    for dept in ("EEE A", "EEE B", "EEE-A", "EEE-B", "EE", "E & E"):
        assert get_report_department(dept) == "EEE"

    for dept in ("ECE A", "ECE B", "ECE-A"):
        assert get_report_department(dept) == "ECE"

    students = pd.DataFrame([
        {
            "S.No": 1, "DEPT": "EEE A", "Department": "EEE A",
            "Report Department": "EEE A", "Name": "A",
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 2, "rating": 1500.0, "ranking": 6000,
            }],
            "level": "Unrated",
        },
        {
            "S.No": 2, "DEPT": "EEE B", "Department": "EEE B",
            "Report Department": "EEE B", "Name": "B",
            "contest_history": [{
                "contest_date": "2026-08-02", "attended": True,
                "problems_solved": 1, "rating": 1400.0, "ranking": 9000,
            }],
            "level": "Unrated",
        },
    ])

    fixed = ensure_student_columns(students)
    assert list(fixed["Report Department"]) == ["EEE", "EEE"]

    report_df, _, _, _, _ = calculate_grouped_department_report(
        fixed, contest_date="2026-08-02"
    )
    eee = report_df[report_df["Dept"] == "EEE"].iloc[0]
    assert eee["Total Strength"] == 2
    assert eee["Total attended"] == 2


def test_generate_missing_data_issues_excel():
    """Verify that Missing Data Excel report contains all expected sheets and issues."""
    from openpyxl import load_workbook
    from services.excel_service import generate_missing_data_issues_excel_bytes

    students = pd.DataFrame([
        {
            "S.No": 1, "Register No": "732124205001", "Name": "Alice",
            "DEPT": "", "Department": "", "Leetcode Link": "https://leetcode.com/u/alice/",
            "fetch_status": "Success",
        },
        {
            "S.No": 2, "Register No": "732124205002", "Name": "Bob",
            "DEPT": "CSE A", "Department": "CSE A", "Leetcode Link": "Long Absent",
            "fetch_status": "Invalid LeetCode Link",
        },
        {
            "S.No": 3, "Register No": "732124205001", "Name": "Duplicate Reg",
            "DEPT": "IT A", "Department": "IT A", "Leetcode Link": "https://leetcode.com/u/charlie/",
            "fetch_status": "Success",
        },
    ])

    buf = generate_missing_data_issues_excel_bytes(students)
    wb = load_workbook(buf)
    sheet_names = wb.sheetnames
    assert "All Data Issues" in sheet_names
    assert "Missing Department" in sheet_names
    assert "Link Issues" in sheet_names
    assert "Duplicate Register Nos" in sheet_names

    ws_dept = wb["Missing Department"]
    assert ws_dept.cell(row=2, column=3).value == "Alice"

    ws_links = wb["Link Issues"]
    assert ws_links.cell(row=2, column=3).value == "Bob"


def test_recent_uploads_tracking():
    """Verify recording and fetching recent uploads is capped at MAX_RECENT_UPLOADS (3)."""
    import app as flask_app
    from config import MAX_RECENT_UPLOADS

    # Test recording 4 uploads
    for i in range(1, 5):
        flask_app._record_recent_upload(
            session_id=f"test_session_{i}",
            filename=f"upload_{i}.xlsx",
            total_records=10 * i,
            contest_date="2026-08-02",
        )

    recent = flask_app._get_recent_uploads()
    assert len(recent) <= MAX_RECENT_UPLOADS
    assert recent[0]["session_id"] == "test_session_4"
    assert recent[0]["filename"] == "upload_4.xlsx"


def test_overall_global_ranking_department_grouping_and_filters():
    """Test CSE (CSE A + CSE B), IT (IT A + IT B), global rank buckets and summaries."""
    from io import BytesIO
    from services.report_service import (
        calculate_overall_ranking_summary,
        calculate_department_wise_global_ranking_comparison,
        filter_weekly_contest_details,
        get_weekly_contest_details,
    )
    from services.excel_service import generate_weekly_contest_details_excel

    students = pd.DataFrame([
        # CSE A
        {"S.No": 1, "Register No": 732124205001, "Name": "Student 1", "DEPT": "CSE A", "Department": "CSE A",
         "contest_ranking": 2087, "excel_contest_attended": 1, "excel_problems_solved": 4, "excel_contest_rating": 2452, "level": "Guardian"},
        # CSE B
        {"S.No": 2, "Register No": 732124205002, "Name": "Student 2", "DEPT": "CSE B", "Department": "CSE B",
         "contest_ranking": 12540, "excel_contest_attended": 1, "excel_problems_solved": 3, "excel_contest_rating": 2180, "level": "Knight"},
        # IT A
        {"S.No": 3, "Register No": 732124205003, "Name": "Student 3", "DEPT": "IT A", "Department": "IT A",
         "contest_ranking": 52340, "excel_contest_attended": 1, "excel_problems_solved": 2, "excel_contest_rating": 1850, "level": "Knight"},
        # IT B
        {"S.No": 4, "Register No": 732124205004, "Name": "Student 4", "DEPT": "IT B", "Department": "IT B",
         "contest_ranking": 125000, "excel_contest_attended": 1, "excel_problems_solved": 1, "excel_contest_rating": 1400, "level": "Unrated"},
        # EEE
        {"S.No": 5, "Register No": 732124205005, "Name": "Student 5", "DEPT": "EEE", "Department": "EEE",
         "contest_ranking": 8500, "excel_contest_attended": 1, "excel_problems_solved": 2, "excel_contest_rating": 1600, "level": "Unrated"},
        # ECE
        {"S.No": 6, "Register No": 732124205006, "Name": "Student 6", "DEPT": "ECE", "Department": "ECE",
         "contest_ranking": None, "excel_contest_attended": 0, "excel_problems_solved": 0, "level": "Unrated"},
        # AI&DS
        {"S.No": 7, "Register No": 732124205007, "Name": "Student 7", "DEPT": "AI&DS", "Department": "AI&DS",
         "contest_ranking": 19500, "excel_contest_attended": 1, "excel_problems_solved": 4, "excel_contest_rating": 2100, "level": "Knight"},
    ])

    # 1. Department Filter CSE groups CSE A and CSE B
    cse_filtered = filter_weekly_contest_details(students, department_filter="CSE")
    assert len(cse_filtered) == 2
    assert set(cse_filtered["Register No"]) == {732124205001, 732124205002}

    # 2. Department Filter IT groups IT A and IT B
    it_filtered = filter_weekly_contest_details(students, department_filter="IT")
    assert len(it_filtered) == 2
    assert set(it_filtered["Register No"]) == {732124205003, 732124205004}

    # 3. Combined Filter: CSE + Below 20,000
    cse_below_20k = filter_weekly_contest_details(students, department_filter="CSE", global_rank_filter="Below 20,000")
    assert len(cse_below_20k) == 2

    # 4. Combined Filter: IT + Above 100,000
    it_above_100k = filter_weekly_contest_details(students, department_filter="IT", global_rank_filter="Above 100,000")
    assert len(it_above_100k) == 1
    assert it_above_100k.iloc[0]["Register No"] == 732124205004

    # 5. Summary calculation for CSE
    cse_summary = calculate_overall_ranking_summary(students, department_filter="CSE")
    assert cse_summary["below_20k"] == 2
    assert cse_summary["from_20k_to_100k"] == 0
    assert cse_summary["above_100k"] == 0
    assert cse_summary["not_available"] == 0
    assert cse_summary["total"] == 2

    # 6. Department comparison across 5 main departments
    comparison = calculate_department_wise_global_ranking_comparison(students)
    dept_names = [row["department"] for row in comparison]
    assert dept_names == ["CSE", "EEE", "ECE", "IT", "AI&DS", "Total"]
    total_row = [r for r in comparison if r["department"] == "Total"][0]
    assert total_row["total"] == 7
    assert total_row["below_20k"] == 4  # 2087, 12540, 8500, 19500
    assert total_row["from_20k_to_100k"] == 1  # 52340
    assert total_row["above_100k"] == 1  # 125000
    assert total_row["not_available"] == 1  # ECE None

    # 7. Excel Export check
    from openpyxl import load_workbook
    buffer = generate_weekly_contest_details_excel(students, output_path=BytesIO())
    assert buffer.getvalue()[:4] == b"PK\x03\x04"
    wb = load_workbook(buffer)
    ws = wb["Weekly Contest"]
    header_row = [cell.value for cell in ws[2]] if ws.max_row >= 2 and ws.cell(row=2, column=1).value == "S.No" else [cell.value for cell in ws[1]]
    assert header_row == [
        "S.No",
        "Register No",
        "Student Name",
        "Department",
        "Contest Rating",
        "Level",
        "Contest Rank",
        "Solved",
        "Overall Global Ranking",
        "Total Problems Solved",
        "Easy",
        "Medium",
        "Hard",
        "Profile Rank",
    ]


def _safe_int(v):
    return int(v) if v is not None else 0


def test_contest_attendance_detection_for_it_and_skewed_records():
    """Verify students in IT with varied attendance flags, solved problems, or rank are marked Attended."""
    import json
    from services.report_service import get_weekly_contest_details

    students = pd.DataFrame([
        {
            "S.No": 1,
            "Register No": "732124205001",
            "Name": "IT Student Attended",
            "DEPT": "IT A",
            "Department": "IT A",
            "contest_history": json.dumps([
                {
                    "title": "Weekly Contest 410",
                    "contest_date": "2026-08-09",
                    "attended": False,  # GraphQL may have false but solved 1 problem
                    "problems_solved": 1,
                    "ranking": 12500,
                    "rating": 1540.5,
                }
            ]),
            "global_ranking": 50000,
        },
        {
            "S.No": 2,
            "Register No": "732124205002",
            "Name": "IT Student Not Attended",
            "DEPT": "IT B",
            "Department": "IT B",
            "contest_history": json.dumps([]),
            "global_ranking": 120000,
        },
    ])

    records, used_date, contest_title, filter_counts = get_weekly_contest_details(
        students,
        contest_date="2026-08-09",
    )
    assert len(records) == 2
    it_attended = next(r for r in records if str(r["Register No"]) == "732124205001")
    it_not_attended = next(r for r in records if str(r["Register No"]) == "732124205002")

    assert it_attended["Contest"] == "Attended"
    assert it_attended["Solved"] == "1/4"
    assert it_not_attended["Contest"] == "Not Attended"
    assert it_not_attended["Solved"] == "-"


def test_attended_and_not_attended_eight_edge_cases():
    """
    Test all 8 specific edge cases from specification:
    CASE 1: Attended + 4/4 -> Attended
    CASE 2: Attended + 3/4 -> Attended
    CASE 3: Attended + 1/4 -> Attended
    CASE 4: Attended + 0/4 -> Attended
    CASE 5: Not attended -> Not Attended
    CASE 6: Not attended + overall profile ranking exists -> Not Attended
    CASE 7: Previous contest attended + selected contest not attended -> Not Attended
    CASE 8: Selected contest attended + previous contest not attended -> Attended
    """
    import json
    from services.report_service import get_weekly_contest_details, build_student_contest_report_record
    from services.leetcode_service import prepare_report_dataframe

    selected_date = "2026-08-09"
    students = pd.DataFrame([
        # CASE 1: Attended + 4/4
        {
            "S.No": 1, "Register No": "732123104001", "Name": "Case 1 Student", "DEPT": "CSE A", "Department": "CSE A",
            "contest_history": json.dumps([{"title": "Weekly Contest 514", "contest_date": "2026-08-09", "attended": True, "problems_solved": 4, "rating": 2150, "ranking": 120}]),
            "global_ranking": 1200, "total_solved": 600,
        },
        # CASE 2: Attended + 3/4
        {
            "S.No": 2, "Register No": "732123104002", "Name": "Case 2 Student", "DEPT": "CSE B", "Department": "CSE B",
            "contest_history": json.dumps([{"title": "Weekly Contest 514", "contest_date": "2026-08-09", "attended": True, "problems_solved": 3, "rating": 1850, "ranking": 2400}]),
            "global_ranking": 18000, "total_solved": 350,
        },
        # CASE 3: Attended + 1/4
        {
            "S.No": 3, "Register No": "732123205001", "Name": "Case 3 Student", "DEPT": "IT A", "Department": "IT A",
            "contest_history": json.dumps([{"title": "Weekly Contest 514", "contest_date": "2026-08-09", "attended": True, "problems_solved": 1, "rating": 1520, "ranking": 12000}]),
            "global_ranking": 85000, "total_solved": 120,
        },
        # CASE 4: Attended + 0/4 (Zero solved)
        {
            "S.No": 4, "Register No": "732123205002", "Name": "Case 4 Student", "DEPT": "IT B", "Department": "IT B",
            "contest_history": json.dumps([{"title": "Weekly Contest 514", "contest_date": "2026-08-09", "attended": True, "problems_solved": 0, "rating": 1450, "ranking": 25000}]),
            "global_ranking": 150000, "total_solved": 40,
        },
        # CASE 5: Not attended
        {
            "S.No": 5, "Register No": "732123106001", "Name": "Case 5 Student", "DEPT": "ECE", "Department": "ECE",
            "contest_history": json.dumps([]),
            "global_ranking": None, "total_solved": 0,
        },
        # CASE 6: Not attended + overall profile ranking exists (e.g. 821,739)
        {
            "S.No": 6, "Register No": "732123106002", "Name": "Case 6 Student", "DEPT": "ECE", "Department": "ECE",
            "contest_history": json.dumps([]),
            "global_ranking": 821739, "total_solved": 444, "profile_ranking": 259135,
        },
        # CASE 7: Previous contest attended (e.g. 2026-08-02) + selected contest not attended (2026-08-09)
        {
            "S.No": 7, "Register No": "732123105001", "Name": "Case 7 Student", "DEPT": "EEE", "Department": "EEE",
            "contest_history": json.dumps([{"title": "Weekly Contest 513", "contest_date": "2026-08-02", "attended": True, "problems_solved": 2, "rating": 1600, "ranking": 4500}]),
            "global_ranking": 60000, "total_solved": 150,
        },
        # CASE 8: Selected contest attended (2026-08-09) + previous contest not attended
        {
            "S.No": 8, "Register No": "732123243001", "Name": "Case 8 Student", "DEPT": "AI&DS", "Department": "AI&DS",
            "contest_history": json.dumps([{"title": "Weekly Contest 514", "contest_date": "2026-08-09", "attended": True, "problems_solved": 2, "rating": 1650, "ranking": 5000}]),
            "global_ranking": 40000, "total_solved": 200,
        },
    ])

    records, used_date, contest_title, filter_counts = get_weekly_contest_details(
        students,
        contest_date=selected_date,
    )
    by_reg = {str(r["Register No"]): r for r in records}

    # Verify CASE 1
    assert by_reg["732123104001"]["Contest"] == "Attended"
    assert by_reg["732123104001"]["Solved"] == "4/4"
    assert by_reg["732123104001"]["Contest Rank"] == 120

    # Verify CASE 2
    assert by_reg["732123104002"]["Contest"] == "Attended"
    assert by_reg["732123104002"]["Solved"] == "3/4"

    # Verify CASE 3
    assert by_reg["732123205001"]["Contest"] == "Attended"
    assert by_reg["732123205001"]["Solved"] == "1/4"

    # Verify CASE 4: Attended + 0/4 (Zero solved) must be Attended
    assert by_reg["732123205002"]["Contest"] == "Attended"
    assert by_reg["732123205002"]["Solved"] == "0/4"
    assert by_reg["732123205002"]["Contest Rank"] == 25000

    # Verify CASE 5: Not attended
    assert by_reg["732123106001"]["Contest"] == "Not Attended"
    assert by_reg["732123106001"]["Solved"] == "-"
    assert by_reg["732123106001"]["Contest Rank"] == "-"

    # Verify CASE 6: Not attended + overall profile ranking exists -> Not Attended
    assert by_reg["732123106002"]["Contest"] == "Not Attended"
    assert by_reg["732123106002"]["Solved"] == "-"
    assert by_reg["732123106002"]["Contest Rank"] == "-"
    assert by_reg["732123106002"]["Overall Global Ranking"] == "821,739"
    assert by_reg["732123106002"]["Profile Rank"] == 259135

    # Verify CASE 7: Previous contest attended + selected contest not attended -> Not Attended
    assert by_reg["732123105001"]["Contest"] == "Not Attended"
    assert by_reg["732123105001"]["Solved"] == "-"
    assert by_reg["732123105001"]["Contest Rank"] == "-"

    # Verify CASE 8: Selected contest attended + previous contest not attended -> Attended
    assert by_reg["732123243001"]["Contest"] == "Attended"
    assert by_reg["732123243001"]["Solved"] == "2/4"

    # Verify Student-Contest-Report export records
    prep_df, _, _ = prepare_report_dataframe(students, contest_date=selected_date)
    sc_records = [build_student_contest_report_record(row) for _, row in prep_df.iterrows()]
    sc_by_reg = {str(r["Register Number"]): r for r in sc_records}

    assert sc_by_reg["732123104001"]["Contest (Attended / Not Attended)"] == "Attended"
    assert sc_by_reg["732123104001"]["Problems Solved (out of 4)"] == "4"
    assert sc_by_reg["732123205002"]["Contest (Attended / Not Attended)"] == "Attended"
    assert sc_by_reg["732123205002"]["Problems Solved (out of 4)"] == "0"
    assert sc_by_reg["732123106002"]["Contest (Attended / Not Attended)"] == "Not Attended"
    assert sc_by_reg["732123106002"]["Problems Solved (out of 4)"] == "-"
    assert sc_by_reg["732123105001"]["Contest (Attended / Not Attended)"] == "Not Attended"
    assert sc_by_reg["732123105001"]["Problems Solved (out of 4)"] == "-"


def test_leetcode_url_with_space_and_special_formats():
    """Verify URLs with embedded spaces (e.g. Suga_ 20), markdown, and trailing spaces extract correct username."""
    from services.leetcode_service import extract_username, clean_leetcode_link

    test_urls = [
        ("https://leetcode.com/u/Suga_ 20/", "Suga_20", "https://leetcode.com/u/Suga_20/"),
        ("https://leetcode.com/u/Suga_20/", "Suga_20", "https://leetcode.com/u/Suga_20/"),
        ("https://leetcode.com/u/Suga_20 /", "Suga_20", "https://leetcode.com/u/Suga_20/"),
        ("https://leetcode.com/u/ Suga_20 /", "Suga_20", "https://leetcode.com/u/Suga_20/"),
        ("[https://leetcode.com/u/Aakil-shihab14/](https://leetcode.com/u/Aakil-shihab14/)", "Aakil-shihab14", "https://leetcode.com/u/Aakil-shihab14/"),
        ("https://leetcode.com/u/_Abisri_/", "_Abisri_", "https://leetcode.com/u/_Abisri_/"),
        ("Suga_20", "Suga_20", "https://leetcode.com/u/Suga_20/"),
    ]

    for raw, exp_user, exp_clean in test_urls:
        assert extract_username(raw) == exp_user, f"Failed username for {raw}: got {extract_username(raw)}"
        assert clean_leetcode_link(raw) == exp_clean, f"Failed clean link for {raw}: got {clean_leetcode_link(raw)}"


if __name__ == "__main__":
    test_classify_problem_solved_contest_range()
    test_grouped_report_problem_buckets_sum_to_strength()
    test_contest_lookup_excel_format()
    test_problem_buckets_only_count_attendees()
    test_rating_and_ranking_boundary_values()
    test_filter_and_sort_students()
    test_all_contest_buckets_sum_to_attended()
    test_s_report_excel_writes_data()
    test_all_report_departments_group_correctly()
    test_eee_department_filter_with_blank_department_column()
    test_eee_section_departments_group_into_eee()
    test_weekly_contest_details_filters_and_records()
    test_student_contest_report_records()
    test_build_language_solved_counts()
    test_generate_missing_data_issues_excel()
    test_recent_uploads_tracking()
    test_overall_global_ranking_department_grouping_and_filters()
    test_contest_attendance_detection_for_it_and_skewed_records()
    test_attended_and_not_attended_eight_edge_cases()
    test_leetcode_url_with_space_and_special_formats()
    print("All tests passed.")




